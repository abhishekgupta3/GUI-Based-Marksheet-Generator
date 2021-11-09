import os
import csv
import openpyxl
from openpyxl import workbook
from openpyxl import styles
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

colors = {
    "RED": '00FF0000',
    "BLUE": '000000FF',
    "GREEN": '00008000'
}
num_questions = 0 # total number of questions
with open('sample_input/responses.csv') as csvFile:
    file = csv.reader(csvFile)
    answers = []
    i = 0
    for row in file:
        if row[0] != 'Timestamp':
            if row[6] == "ANSWER":
                size = len(row)
                num_questions = size - 7 

                for itr in range(1, size - 6):
                    answers.append(row[itr + 6])

            wb = openpyxl.Workbook() # create a new workbook
            sheet = wb.active
            sheet.title = 'quiz'
            
            sheet.column_dimensions['A'].width = 17.7
            sheet.column_dimensions['B'].width = 17.7
            sheet.column_dimensions['C'].width = 17.7
            sheet.column_dimensions['D'].width = 17.7
            sheet.column_dimensions['E'].width = 17.7

            img = openpyxl.drawing.image.Image('iitp_logo.png') # adding image in excel
            img.anchor = 'A1'
            img.height = 81
            img.width = 630
            sheet.add_image(img)

            sheet.merge_cells('A5:E5')

            sheet['A5'].value = 'Mark Sheet'
            sheet['A5'].font = Font(name = 'Century', size = 18, bold = True, underline = 'single')
            sheet['A5'].alignment = Alignment(horizontal = 'center', vertical = 'bottom')
            sheet['A5'].border = Border()

            sheet['A6'] = 'Name:'
            sheet['A6'].font = Font(name = 'Century', size = 12)
            sheet['A6'].alignment = Alignment(horizontal = 'right')

            sheet['B6'] = row[3]
            sheet['B6'].font = Font(name = 'Century', size = 12, bold = True)
            sheet['B6'].alignment = Alignment(horizontal = 'left')

            sheet['D6'] = 'Exam:'
            sheet['D6'].font = Font(name = 'Century', size = 12)
            sheet['D6'].alignment = Alignment(horizontal = 'right')

            sheet['E6'] = 'quiz'
            sheet['E6'].font = Font(name = 'Century', size = 12, bold = True)
            sheet['E6'].alignment = Alignment(horizontal = 'left')

            sheet['A7'] = 'Roll Number:'
            sheet['A7'].font = Font(name = 'Century', size = 12)
            sheet['A7'].alignment = Alignment(horizontal = 'right')

            sheet['B7'] = row[6]
            sheet['B7'].font = Font(name = 'Century', size = 12, bold = True)
            sheet['B7'].alignment = Alignment(horizontal = 'left')

            sheet.append(['', '', '', '', ''])

            sheet['B9'] = 'Right'
            sheet['B9'].font = Font(name = 'Century', size = 12, bold = True)
            sheet['B9'].alignment = Alignment(horizontal = 'center')

            sheet['C9'] = 'Wrong'
            sheet['C9'].font = Font(name = 'Century', size = 12, bold = True)
            sheet['C9'].alignment = Alignment(horizontal = 'center')

            sheet['D9'] = 'Not Attempt'
            sheet['D9'].font = Font(name = 'Century', size = 12, bold = True)
            sheet['D9'].alignment = Alignment(horizontal = 'center')

            sheet['E9'] = 'Max'
            sheet['E9'].font = Font(name = 'Century', size = 12, bold = True)
            sheet['E9'].alignment = Alignment(horizontal = 'center')

            sheet['A10'] = 'No.'
            sheet['A10'].font = Font(name = 'Century', size = 12, bold = True)
            sheet['A10'].alignment = Alignment(horizontal = 'center')

            sheet['A11'] = 'Marking'
            sheet['A11'].font = Font(name = 'Century', size = 12, bold = True)
            sheet['A11'].alignment = Alignment(horizontal = 'center')

            sheet['A12'] = 'Total'
            sheet['A12'].font = Font(name = 'Century', size = 12, bold = True)
            sheet['A12'].alignment = Alignment(horizontal = 'center')

            for r in sheet.iter_rows(min_row = 9, min_col = 1, max_row = 12, max_col = 5):
                for cell in r:
                    cell.border = Border(left = Side(style='thin'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))
                    cell.alignment = Alignment(horizontal = 'center')
                    cell.font = Font(name = 'Century', size = 12)
 
            sheet['E10'] = num_questions
            sheet.append(['', '', '', '', ''])
            sheet.append(['', '', '', '', ''])

            sheet['A15'] = 'Student Ans' 
            sheet['A15'].font = Font(name = 'Century', size = 12, bold = True)
            sheet['A15'].alignment = Alignment(horizontal = 'center')
            sheet['A15'].border = Border(left = Side(style='thin'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))

            sheet['B15'] = 'Correct Ans' 
            sheet['B15'].font = Font(name = 'Century', size = 12, bold = True)
            sheet['B15'].alignment = Alignment(horizontal = 'center')
            sheet['B15'].border = Border(left = Side(style='thin'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))

            correct_ans = 0
            wrong_ans = 0

            for itr in range(min(25, num_questions)):
                cell_pos = 'A' + str(itr + 16)
                sheet[cell_pos] = row[7 + itr]
                sheet[cell_pos].alignment = Alignment(horizontal = 'center')
                if row[7 + itr] == answers[itr]:
                    sheet[cell_pos].font = Font(name = 'Century', size = 12, color = colors["GREEN"])
                    correct_ans += 1
                else:
                    sheet[cell_pos].font = Font(name = 'Century', size = 12, color = colors["RED"])
                    if len(sheet[cell_pos].value) > 0:
                        wrong_ans += 1
                sheet[cell_pos].border = Border(left = Side(style='thin'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))

                cell_pos = 'B' + str(itr + 16)
                sheet[cell_pos] = answers[itr]
                sheet[cell_pos].alignment = Alignment(horizontal = 'center')
                sheet[cell_pos].font = Font(name = 'Century', size = 12, color = colors["BLUE"])
                sheet[cell_pos].border = Border(left = Side(style='thin'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))

            remaining_ques = num_questions - 25

            if remaining_ques:
                sheet['D15'] = 'Student Ans' 
                sheet['D15'].font = Font(name = 'Century', size = 12, bold = True)
                sheet['D15'].alignment = Alignment(horizontal = 'center')
                sheet['D15'].border = Border(left = Side(style='thin'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))

                sheet['E15'] = 'Correct Ans' 
                sheet['E15'].font = Font(name = 'Century', size = 12, bold = True)
                sheet['E15'].alignment = Alignment(horizontal = 'center')
                sheet['E15'].border = Border(left = Side(style='thin'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))

                for itr in range(remaining_ques):
                    cell_pos = 'D' + str(itr + 16)
                    sheet[cell_pos] = row[32 + itr]
                    sheet[cell_pos].alignment = Alignment(horizontal = 'center')
                    if row[32 + itr] == answers[25 + itr]:
                        sheet[cell_pos].font = Font(name = 'Century', size = 12, color = colors["GREEN"])
                        correct_ans += 1
                    else:
                        sheet[cell_pos].font = Font(name = 'Century', size = 12, color = colors["RED"])
                        if len(sheet[cell_pos].value) > 0:
                            wrong_ans += 1
                    sheet[cell_pos].border = Border(left = Side(style='thin'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))

                    cell_pos = 'E' + str(itr + 16)
                    sheet[cell_pos] = answers[25 + itr]
                    sheet[cell_pos].alignment = Alignment(horizontal = 'center')
                    sheet[cell_pos].font = Font(name = 'Century', size = 12, color = colors["BLUE"])
                    sheet[cell_pos].border = Border(left = Side(style='thin'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))

            
            sheet['B10'] = correct_ans
            sheet['B10'].font = Font(name = 'Century', size = 12, color = colors["GREEN"])
            
            sheet['B11'] = 5 # user input of postive marks for correct ans
            sheet['B11'].font = Font(name = 'Century', size = 12, color = colors["GREEN"])

            sheet['B12'] = sheet['B11'].value * sheet['B10'].value
            sheet['B12'].font = Font(name = 'Century', size = 12, color = colors["GREEN"])

            sheet['C10'] = wrong_ans
            sheet['C10'].font = Font(name = 'Century', size = 12, color = colors["RED"])

            sheet['C11'] = -1 # user input of negative marks for wrong ans
            sheet['C11'].font = Font(name = 'Century', size = 12, color = colors["RED"])

            sheet['C12'] = sheet['C11'].value * sheet['C10'].value
            sheet['C12'].font = Font(name = 'Century', size = 12, color = colors["RED"])

            sheet['D10'] = num_questions - correct_ans - wrong_ans
            sheet['D11'] = 0
            
            total_marks = sheet['B11'].value * num_questions
            student_marks = sheet['B12'].value + sheet['C12'].value

            sheet['E12'] = f'{student_marks}/{total_marks}'
            sheet['E12'].font = Font(name = 'Century', size = 12, color = colors["BLUE"])

            wb.save(f'output/{row[6]}.xlsx')
            i += 1
            if i > 10:
                break
