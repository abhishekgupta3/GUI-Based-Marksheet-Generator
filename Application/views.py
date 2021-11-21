from django.http import response
from django.shortcuts import render, HttpResponse
from . models import myuploadfile
import os
import csv
import openpyxl
from openpyxl import workbook, load_workbook
from openpyxl import styles
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from django.core.mail import EmailMessage

colors = {
    "RED": '00FF0000',
    "BLUE": '000000FF',
    "GREEN": '00008000'
}
answers = [] # correct answers
num_questions = 0 # total number of questions

def create_excel(row, postive_marks, negative_marks, size, num_questions):
    wb = openpyxl.Workbook() # create a new workbook
    sheet = wb.active
    sheet.title = 'quiz'
    
    sheet.column_dimensions['A'].width = 17.7
    sheet.column_dimensions['B'].width = 17.7
    sheet.column_dimensions['C'].width = 17.7
    sheet.column_dimensions['D'].width = 17.7
    sheet.column_dimensions['E'].width = 17.7

    img = openpyxl.drawing.image.Image('static/images/iitp_logo.png') # adding image in excel
    img.anchor = 'A1'
    img.height = 81
    img.width = 630
    sheet.add_image(img)

    sheet.merge_cells('A5:E5')

    sheet['A5'].value = 'Mark Sheet'
    sheet['A5'].font = Font(name = 'Century', size = 18, bold = True, underline = 'single')
    sheet['A5'].alignment = Alignment(horizontal = 'center', vertical = 'bottom')
    sheet['A5'].border = Border()

    sheet['A6'] = 'Name:'
    sheet['A6'].font = Font(name = 'Century', size = 12)
    sheet['A6'].alignment = Alignment(horizontal = 'right')

    sheet['B6'] = row[3]
    sheet['B6'].font = Font(name = 'Century', size = 12, bold = True)
    sheet['B6'].alignment = Alignment(horizontal = 'left')

    sheet['D6'] = 'Exam:'
    sheet['D6'].font = Font(name = 'Century', size = 12)
    sheet['D6'].alignment = Alignment(horizontal = 'right')

    sheet['E6'] = 'quiz'
    sheet['E6'].font = Font(name = 'Century', size = 12, bold = True)
    sheet['E6'].alignment = Alignment(horizontal = 'left')

    sheet['A7'] = 'Roll Number:'
    sheet['A7'].font = Font(name = 'Century', size = 12)
    sheet['A7'].alignment = Alignment(horizontal = 'right')

    roll_no = row[6].upper()
    sheet['B7'] = roll_no
    sheet['B7'].font = Font(name = 'Century', size = 12, bold = True)
    sheet['B7'].alignment = Alignment(horizontal = 'left')

    sheet.append(['', '', '', '', ''])

    sheet['B9'] = 'Right'
    sheet['B9'].font = Font(name = 'Century', size = 12, bold = True)
    sheet['B9'].alignment = Alignment(horizontal = 'center')

    sheet['C9'] = 'Wrong'
    sheet['C9'].font = Font(name = 'Century', size = 12, bold = True)
    sheet['C9'].alignment = Alignment(horizontal = 'center')

    sheet['D9'] = 'Not Attempt'
    sheet['D9'].font = Font(name = 'Century', size = 12, bold = True)
    sheet['D9'].alignment = Alignment(horizontal = 'center')

    sheet['E9'] = 'Max'
    sheet['E9'].font = Font(name = 'Century', size = 12, bold = True)
    sheet['E9'].alignment = Alignment(horizontal = 'center')

    sheet['A10'] = 'No.'
    sheet['A10'].font = Font(name = 'Century', size = 12, bold = True)
    sheet['A10'].alignment = Alignment(horizontal = 'center')

    sheet['A11'] = 'Marking'
    sheet['A11'].font = Font(name = 'Century', size = 12, bold = True)
    sheet['A11'].alignment = Alignment(horizontal = 'center')

    sheet['A12'] = 'Total'
    sheet['A12'].font = Font(name = 'Century', size = 12, bold = True)
    sheet['A12'].alignment = Alignment(horizontal = 'center')

    for r in sheet.iter_rows(min_row = 9, min_col = 1, max_row = 12, max_col = 5):
        for cell in r:
            cell.border = Border(left = Side(style='thin'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))
            cell.alignment = Alignment(horizontal = 'center')
            cell.font = Font(name = 'Century', size = 12)

    sheet['E10'] = num_questions
    sheet.append(['', '', '', '', ''])
    sheet.append(['', '', '', '', ''])

    sheet['A15'] = 'Student Ans' 
    sheet['A15'].font = Font(name = 'Century', size = 12, bold = True)
    sheet['A15'].alignment = Alignment(horizontal = 'center')
    sheet['A15'].border = Border(left = Side(style='thin'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))

    sheet['B15'] = 'Correct Ans' 
    sheet['B15'].font = Font(name = 'Century', size = 12, bold = True)
    sheet['B15'].alignment = Alignment(horizontal = 'center')
    sheet['B15'].border = Border(left = Side(style='thin'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))

    correct_ans = 0
    wrong_ans = 0

    for itr in range(min(25, num_questions)):
        cell_pos = 'A' + str(itr + 16)
        sheet[cell_pos] = row[7 + itr]
        sheet[cell_pos].alignment = Alignment(horizontal = 'center')
        if row[7 + itr] == answers[itr]:
            sheet[cell_pos].font = Font(name = 'Century', size = 12, color = colors["GREEN"])
            correct_ans += 1
        else:
            sheet[cell_pos].font = Font(name = 'Century', size = 12, color = colors["RED"])
            if len(sheet[cell_pos].value) > 0:
                wrong_ans += 1
        sheet[cell_pos].border = Border(left = Side(style='thin'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))

        cell_pos = 'B' + str(itr + 16)
        sheet[cell_pos] = answers[itr]
        sheet[cell_pos].alignment = Alignment(horizontal = 'center')
        sheet[cell_pos].font = Font(name = 'Century', size = 12, color = colors["BLUE"])
        sheet[cell_pos].border = Border(left = Side(style='thin'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))

    remaining_ques = num_questions - 25

    if remaining_ques:
        sheet['D15'] = 'Student Ans' 
        sheet['D15'].font = Font(name = 'Century', size = 12, bold = True)
        sheet['D15'].alignment = Alignment(horizontal = 'center')
        sheet['D15'].border = Border(left = Side(style='thin'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))

        sheet['E15'] = 'Correct Ans' 
        sheet['E15'].font = Font(name = 'Century', size = 12, bold = True)
        sheet['E15'].alignment = Alignment(horizontal = 'center')
        sheet['E15'].border = Border(left = Side(style='thin'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))

        for itr in range(remaining_ques):
            cell_pos = 'D' + str(itr + 16)
            sheet[cell_pos] = row[32 + itr]
            sheet[cell_pos].alignment = Alignment(horizontal = 'center')
            if row[32 + itr] == answers[25 + itr]:
                sheet[cell_pos].font = Font(name = 'Century', size = 12, color = colors["GREEN"])
                correct_ans += 1
            else:
                sheet[cell_pos].font = Font(name = 'Century', size = 12, color = colors["RED"])
                if len(sheet[cell_pos].value) > 0:
                    wrong_ans += 1
            sheet[cell_pos].border = Border(left = Side(style='thin'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))

            cell_pos = 'E' + str(itr + 16)
            sheet[cell_pos] = answers[25 + itr]
            sheet[cell_pos].alignment = Alignment(horizontal = 'center')
            sheet[cell_pos].font = Font(name = 'Century', size = 12, color = colors["BLUE"])
            sheet[cell_pos].border = Border(left = Side(style='thin'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))
    
    sheet['B10'] = correct_ans
    sheet['B10'].font = Font(name = 'Century', size = 12, color = colors["GREEN"])
    
    sheet['B11'] = int(postive_marks) # user input of postive marks for correct ans
    sheet['B11'].font = Font(name = 'Century', size = 12, color = colors["GREEN"])

    sheet['B12'] = int(sheet['B11'].value) * int(sheet['B10'].value)
    sheet['B12'].font = Font(name = 'Century', size = 12, color = colors["GREEN"])

    sheet['C10'] = int(wrong_ans)
    sheet['C10'].font = Font(name = 'Century', size = 12, color = colors["RED"])

    sheet['C11'] = int(negative_marks) # user input of negative marks for wrong ans
    sheet['C11'].font = Font(name = 'Century', size = 12, color = colors["RED"])

    sheet['C12'] = int(sheet['C11'].value) * int(sheet['C10'].value)
    sheet['C12'].font = Font(name = 'Century', size = 12, color = colors["RED"])

    sheet['D10'] = num_questions - correct_ans - wrong_ans
    sheet['D11'] = 0
    
    total_marks = int(sheet['B11'].value) * int(num_questions)
    student_marks = int(sheet['B12'].value) + int(sheet['C12'].value)

    final_score = f'{student_marks}/{total_marks}'
    sheet['E12'] = final_score
    sheet['E12'].font = Font(name = 'Century', size = 12, color = colors["BLUE"])

    wb.save(f'output/marksheet/{roll_no}.xlsx')
    
# Create your views here.
def index(request):
    context = {}
    if os.path.exists('media'):
        dir = 'media'
        for f in os.listdir(dir):
            os.remove(os.path.join(dir, f))

    if os.path.exists('output/marksheet'):
        dir = 'output/marksheet'
        for f in os.listdir(dir):
            os.remove(os.path.join(dir, f))
        
    return render(request, "index.html", context)

def roll_marksheet(request):
    if request.method == "POST" :
        file1 = request.FILES.getlist("master_roll")
        file2 = request.FILES.getlist("responses")
        postive_marks = request.POST.get("postive_marks")
        negative_marks = request.POST.get("negative_marks")

        master_roll = file1[0] 
        responses = file2[0]
        myuploadfile(master_roll = master_roll, responses = responses, postive_marks = postive_marks, negative_marks = negative_marks).save()

        # ANSWER row validation
        with open('media/responses.csv') as csvFile:
            file = csv.reader(csvFile)
            for row in file:
                if row[0] != 'Timestamp':
                    if row[6] == "ANSWER":
                        size = len(row)
                        num_questions = size - 7 

                        for itr in range(1, size - 6):
                            answers.append(row[itr + 6])
                        break

        if len(answers) == 0:
            return HttpResponse("no roll number with ANSWER is present, Cannot Process!")

        # Create Excel sheet of responses
        with open('media/responses.csv') as csvFile:
            file = csv.reader(csvFile)
            for row in file:
                if row[0] != 'Timestamp':
                    size = len(row)
                    num_questions = size - 7 
                    create_excel(row, postive_marks, negative_marks, size, num_questions)
                else:
                    if not os.path.exists('output'):
                        os.makedirs('output')

                    if not os.path.exists('output/marksheet'):
                        os.makedirs('output/marksheet')         

        # Create Excel sheet of absentees
        with open('media/master_roll.csv') as csvFile:
            file = csv.reader(csvFile)
            for row in file:
                if row[0] != 'roll':
                    roll_no = row[0].upper()
                    if not os.path.exists(f'output/marksheet/{roll_no}.xlsx'):

                        # create blank sheet
                        wb = openpyxl.Workbook() # create a new workbook
                        sheet = wb.active
                        sheet.title = 'quiz'
                        wb.save(f'output/marksheet/{roll_no}.xlsx')

        return render(request, "index.html")

def concise_marksheet(request):
    stud_present = set() # set of roll number given the quiz

    with open('media/responses.csv') as csvFile:
        file = csv.reader(csvFile)
        for row in file:
            if row[0] != 'Timestamp':
                roll_no = row[6].upper()
                if row[6] == "ANSWER":
                    size = len(row)
                    num_questions = size - 7 
                stud_present.add(roll_no)
    
    with open('media/responses.csv') as csvFile:
        file = csv.reader(csvFile)
        for row in file:
            if row[0] != 'Timestamp':
                size = len(row)
                roll_no = row[6].upper()
                num_questions = size - 7 

                wb = load_workbook(f'output/marksheet/{roll_no}.xlsx')
                sheet = wb.active
                final_score = sheet['E12'].value
                correct_ans = sheet['B10'].value
                wrong_ans = sheet['C10'].value
                temp_row = row
                temp_row.insert(6, final_score)
                statusAns = f'[{correct_ans},{wrong_ans},{num_questions - correct_ans - wrong_ans}]'
                temp_row.append(statusAns)

                try:
                    with open('output/marksheet/concise_marksheet.csv', 'a', newline='') as csvFile:
                        writer = csv.writer(csvFile)
                        writer.writerow(temp_row)
                        csvFile.close()
                except:
                    return HttpResponse('Some Error Occured...')
            else:
                # create concise marksheet
                with open('output/marksheet/concise_marksheet.csv', 'w', newline='') as csvFile:
                    writer = csv.writer(csvFile)
                    temp_row = row
                    temp_row.insert(6, 'Score_After_Negative')

                    cur_size = len(temp_row)
                    temp_row.insert(cur_size + num_questions, 'statusAns')
                    
                    for iter in range(len(temp_row)):
                        if len(temp_row[iter]) == 0:
                            temp_row[iter] = f'Unnamed: {iter}'
                        
                    writer.writerow(temp_row)
                    csvFile.close()

        # absentees
        with open('media/master_roll.csv') as csvFile:
            file = csv.reader(csvFile)
            for row in file:
                if row[0] != 'roll':
                    roll_no = row[0].upper()
                    if roll_no not in stud_present:
                        temp_row = ['', '', 'Absent', row[1], '', '', 'Absent', roll_no]
                            
                        with open('output/marksheet/concise_marksheet.csv', 'a', newline='') as csvFile:
                            writer = csv.writer(csvFile)
                            writer.writerow(temp_row)
                            csvFile.close() 


    return render(request, 'index.html')

# Send Email
def send_email(request):
    try:
        with open('media/responses.csv') as csvFile:
            file = csv.reader(csvFile)
            for row in file:
                if row[0] != 'Timestamp':
                    message = EmailMessage('Quiz Marksheet', '', to=[row[1], row[4]])
                    roll_no = row[6].upper()
                    message.attach_file(f'output/marksheet/{roll_no}.xlsx')
                    message.send()
    except:
        return HttpResponse('Some Error Occured...')

    return HttpResponse('Email Sent Successfully...')